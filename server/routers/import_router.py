import csv
import io
import json
import math
import os
import sqlite3 as _sqlite3
import tempfile
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from database import get_connection
from auth import get_current_user
from limiter import user_limiter


def _get_ai_categorizer():
    """
    Returns ai_suggest_category if OPENAI_API_KEY is set, else None.
    Lazy-checked at call time so the router still loads when AI deps are absent.
    """
    if not os.environ.get("OPENAI_API_KEY"):
        return None
    try:
        from ai_helpers import ai_suggest_category
        return ai_suggest_category
    except ImportError:
        return None


def _is_other_fallback(expense_type, valid_types):
    """True when _infer_type returned the catch-all 'Other' — no keyword matched."""
    for v in valid_types:
        if v.lower() == "other" and v == expense_type:
            return True
    return False

router = APIRouter()

HEADER_KEYWORDS = {
    "date", "amount", "description", "name", "total", "balance",
    "debit", "credit", "memo", "category", "type", "note", "notes",
    "payee", "merchant", "transaction", "reference",
}


def _cell_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _parse_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    return [[cell.strip() for cell in row] for row in reader]


def _xlsx_sheet_names(content: bytes) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _parse_xlsx(content: bytes, sheet_name: Optional[str] = None) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append([_cell_str(cell.value) for cell in row])
    wb.close()
    return rows


def _parse_pdf(content: bytes) -> list[list[str]]:
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF support not available on this server.")

    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            all_rows: list[list[str]] = []
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    for row in table:
                        all_rows.append([str(c).strip() if c is not None else "" for c in row])
            if all_rows:
                return all_rows
            full_text = "\n".join(p.extract_text() or "" for p in pdf.pages).strip()
    except HTTPException:
        raise
    except Exception as exc:
        s = f"{type(exc).__name__} {exc}".lower()
        if "password" in s or "encrypt" in s:
            raise HTTPException(status_code=422, detail="This PDF is password-protected. Please unlock it before uploading.")
        raise HTTPException(status_code=422, detail=f"Could not read PDF: {exc}")

    if not full_text:
        raise HTTPException(
            status_code=422,
            detail="This PDF appears to be scanned or image-only. Please upload a text-based bank statement.",
        )

    if os.environ.get("OPENAI_API_KEY"):
        try:
            from openai import OpenAI
            client = OpenAI()
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                response_format={"type": "json_object"},
                temperature=0,
                max_tokens=2000,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Extract all transactions from this bank statement text. "
                            "Return JSON: {\"rows\": [[\"Date\",\"Description\",\"Amount\"], ...]}. "
                            "First row must be column headers. Use negative amounts for debits/charges, "
                            "positive for credits/deposits. Include only transaction rows — no totals, "
                            "summaries, or opening/closing balance lines."
                        ),
                    },
                    {"role": "user", "content": full_text[:8000]},
                ],
            )
            data = json.loads(resp.choices[0].message.content)
            rows = [[str(c) for c in row] for row in data.get("rows", [])]
            if rows:
                return rows
        except Exception:
            pass

    raise HTTPException(
        status_code=422,
        detail="Could not extract a transaction table from this PDF. Try exporting as CSV instead.",
    )


def _detect_header_row(rows: list[list[str]]) -> int:
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows):
        score = 0
        for cell in row:
            lower = cell.lower().replace(" ", "").replace("_", "").replace(".", "")
            if lower in HEADER_KEYWORDS:
                score += 3
            elif cell and not _is_numeric(cell):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _is_numeric(s: str) -> bool:
    try:
        float(s.replace(",", "").replace("$", "").replace("-", "").replace("(", "").replace(")", ""))
        return True
    except ValueError:
        return bool(not s)


def _rows_to_dicts(rows: list[list[str]], header_row: int) -> tuple[list[str], list[dict]]:
    if header_row >= len(rows):
        return [], []
    headers = rows[header_row]
    data_rows = rows[header_row + 1:]
    dicts = []
    for row in data_rows:
        if not any(cell for cell in row):
            continue
        padded = row + [""] * (len(headers) - len(row))
        dicts.append({headers[i]: padded[i] for i in range(len(headers))})
    return headers, dicts


def _parse_date(raw: str) -> str:
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Invalid date: '{raw}'. Accepted formats: YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY")


def _soft_match_type(raw: str, valid: list[str]) -> Optional[str]:
    for v in valid:
        if v.lower() == raw.lower():
            return v
    return None


KEYWORDS = {
    "food & drink": ["restaurant", "food", "grocery", "groceries", "cafe", "coffee", "lunch", "dinner",
                     "breakfast", "pizza", "burger", "sushi", "diner", "bakery", "donut", "sandwich"],
    "transportation": ["uber", "lyft", "gas", "fuel", "parking", "transit", "metro", "bus", "train",
                        "airline", "flight", "taxi", "toll", "shell", "chevron", "bp", "car"],
    "rent & utilities": ["rent", "mortgage", "electric", "electricity", "water", "internet", "cable",
                          "utilities", "utility", "hoa", "insurance"],
    "medical": ["pharmacy", "cvs", "walgreens", "doctor", "hospital", "medical", "dental",
                "health", "gym", "fitness", "clinic", "optician"],
    "entertainment": ["netflix", "spotify", "hulu", "disney", "apple tv", "movie", "cinema",
                      "theater", "concert", "game", "steam", "xbox", "playstation"],
    "shopping": ["amazon", "walmart", "target", "costco", "store", "shop", "mall", "ebay"],
}


def _infer_type_with_source(name: str, valid_types: list[str], rules: list[tuple] = []) -> tuple[str, str]:
    """Returns (category, source) where source is 'rule', 'keyword', or 'fallback'."""
    name_lower = name.lower()
    for pattern, expense_type in rules:
        if pattern.lower() in name_lower:
            matched = _soft_match_type(expense_type, valid_types)
            if matched:
                return matched, "rule"
    for category, keywords in KEYWORDS.items():
        if any(kw in name_lower for kw in keywords):
            matched = _soft_match_type(category, valid_types)
            if not matched:
                matched = next((v for v in valid_types if category in v.lower()), None)
            if matched:
                return matched, "keyword"
    fallback = next((v for v in valid_types if v.lower() == "other"), None) or (valid_types[0] if valid_types else "Other")
    return fallback, "fallback"


def _infer_type(name: str, valid_types: list[str], rules: list[tuple] = []) -> str:
    category, _ = _infer_type_with_source(name, valid_types, rules)
    return category


INCOME_SIGNALS = {"credit", "cr", "income", "deposit", "refund", "payment received"}
EXPENSE_SIGNALS = {"debit", "dr", "expense", "withdrawal", "purchase", "charge"}


def _determine_record_type(amount_raw: float, record_type_col_val: str) -> str:
    if record_type_col_val:
        val = record_type_col_val.strip().lower()
        if any(s in val for s in INCOME_SIGNALS):
            return "income"
        if any(s in val for s in EXPENSE_SIGNALS):
            return "expense"
    return "expense" if amount_raw < 0 else "income"


_MAX_IMPORT_FILE_SIZE = 25 * 1024 * 1024


def _check_file_size(content: bytes):
    if len(content) > _MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Please use a file under 25MB.")


def _get_raw_rows(content: bytes, filename: str, sheet_name: Optional[str] = None) -> list[list[str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    elif lower.endswith((".xlsx", ".xls")):
        return _parse_xlsx(content, sheet_name)
    elif lower.endswith(".pdf"):
        return _parse_pdf(content)
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Upload a .csv, .xlsx, or .pdf file.")


@router.get("/import/infer-type")
def infer_type_endpoint(
    name: str = Query(..., description="Transaction name/description to categorize"),
    user_id: str = Depends(get_current_user),
):
    with get_connection() as conn:
        valid_types = [r["name"] for r in conn.execute(
            "SELECT name FROM expense_types WHERE user_id = %s ORDER BY sort_order", (user_id,)
        ).fetchall()]
        rules = conn.execute(
            "SELECT pattern, expense_type FROM import_rules WHERE user_id = %s", (user_id,)
        ).fetchall()
    if not valid_types:
        return {"type": "Other"}
    return {"type": _infer_type(name, valid_types, rules)}


@router.post("/import/suggest")
@user_limiter.limit("10/hour")
async def suggest_import_categories(
    request: Request,
    file: UploadFile = File(...),
    mapping: str = Form(...),
    header_row: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    user_id: str = Depends(get_current_user),
):
    """Run the full inference pipeline without writing to DB. Returns per-row category suggestions."""
    try:
        col_map = json.loads(mapping)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")

    content = await file.read()
    _check_file_size(content)
    raw_rows = _get_raw_rows(content, file.filename, sheet_name)
    _, data_rows = _rows_to_dicts(raw_rows, header_row)

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM expense_types WHERE user_id = %s ORDER BY sort_order", (user_id,))
    valid_types = [r["name"] for r in cursor.fetchall()]
    cursor.execute("SELECT pattern, expense_type FROM import_rules WHERE user_id = %s", (user_id,))
    rules = [(r["pattern"], r["expense_type"]) for r in cursor.fetchall()]
    from database import get_user_settings
    user_ai_settings = get_user_settings(conn, user_id)
    conn.close()

    ai_categorize_fn = _get_ai_categorizer() if user_ai_settings["ai_enabled"] else None

    AI_CAP = 50
    ai_calls = 0
    ai_rows_count = 0
    income_count = 0
    rows = []

    for i, row in enumerate(data_rows):
        try:
            name = row.get(col_map.get("name", ""), "").strip()
            if not name:
                continue
            raw_amt = row.get(col_map.get("amount", ""), "")
            amount_float = float(str(raw_amt).replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip())
            # float("nan")/float("inf") parse without raising ValueError, and NaN
            # fails every comparison (including == 0), so both would otherwise slip
            # past the zero-check below and corrupt SUM() aggregates downstream.
            if not math.isfinite(amount_float) or amount_float == 0 or abs(amount_float) > 100_000_000:
                continue
            amount = round(abs(amount_float), 2)
            raw_date = row.get(col_map.get("date", ""), "").strip()
            date_str = _parse_date(raw_date)
            record_type_col_val = row.get(col_map.get("record_type", ""), "") if col_map.get("record_type") else ""
            row_type = _determine_record_type(amount_float, record_type_col_val)

            if row_type == "income":
                income_count += 1
                continue

            raw_type = row.get(col_map.get("type", ""), "").strip() if col_map.get("type") else ""
            ai_reasoning = None

            if raw_type and raw_type.lower() == "savings":
                suggested_type = "Savings"
                source = "file"
            elif raw_type:
                matched = _soft_match_type(raw_type, valid_types)
                if matched:
                    suggested_type, source = matched, "file"
                else:
                    suggested_type, source = _infer_type_with_source(name, valid_types, rules)
            else:
                suggested_type, source = _infer_type_with_source(name, valid_types, rules)

            if source == "fallback" and ai_categorize_fn and ai_calls < AI_CAP:
                suggestion = ai_categorize_fn(name, amount, valid_types)
                ai_calls += 1
                suggested = suggestion.get("category")
                if suggested and suggested in valid_types:
                    suggested_type = suggested
                    source = "ai"
                    ai_reasoning = suggestion.get("reasoning")
                    ai_rows_count += 1

            rows.append({
                "row_idx": i,
                "name": name,
                "amount": amount,
                "date": date_str,
                "record_type": "expense",
                "suggested_type": suggested_type,
                "source": source,
                "ai_reasoning": ai_reasoning,
            })
        except (ValueError, KeyError, TypeError):
            continue

    return {
        "rows": rows,
        "income_count": income_count,
        "ai_rows_count": ai_rows_count,
        "ai_cap_reached": ai_calls >= AI_CAP and ai_calls > 0,
        "ai_enabled": user_ai_settings["ai_enabled"],
        "valid_types": valid_types,
    }


_DEFAULT_COLORS = [
    "#e8a87c", "#82b4e0", "#c49ee8", "#f0c040", "#80cbc4", "#a0a0a0",
    "#ef9a9a", "#90caf9", "#a5d6a7", "#ffcc80", "#ce93d8", "#f48fb1",
    "#ff8a65", "#4db6ac", "#7986cb", "#aed581",
]
_DEFAULT_ICON = "Category"


@router.post("/import/budgets")
async def import_budgets(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    header_row: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    target_month: Optional[str] = Form(None),
    user_id: str = Depends(get_current_user),
):
    mapping_dict = json.loads(mapping)
    cat_col = mapping_dict.get("category", "")
    limit_col = mapping_dict.get("monthly_limit", "")
    month_col = mapping_dict.get("month", "")

    content = await file.read()
    _check_file_size(content)
    raw_rows = _get_raw_rows(content, file.filename, sheet_name)
    _, data_rows = _rows_to_dicts(raw_rows, header_row)

    aggregated: dict[tuple, float] = {}
    display_name: dict[str, str] = {}
    errors = []
    skipped = 0

    for i, row in enumerate(data_rows, start=header_row + 2):
        cat_raw = row.get(cat_col, "").strip()
        limit_raw = row.get(limit_col, "").strip().replace(",", "").lstrip("$").lstrip("-").strip()
        month_raw = row.get(month_col, "").strip() if month_col else ""

        if not cat_raw:
            errors.append({"row": i, "reason": "Missing category"})
            skipped += 1
            continue

        try:
            limit_val = float(limit_raw)
            # float("nan")/float("inf") parse without raising — NaN in particular
            # would poison every later row's running total once summed together
            # in `aggregated`, not just this one row.
            if not math.isfinite(limit_val) or limit_val > 100_000_000:
                raise ValueError("out of range")
        except ValueError:
            errors.append({"row": i, "reason": f"Invalid amount: '{row.get(limit_col, '')}'"})
            skipped += 1
            continue

        if target_month:
            month_key = target_month
        elif month_raw:
            try:
                month_key = _parse_date(month_raw)[:7]
            except ValueError:
                errors.append({"row": i, "reason": f"Invalid month: '{month_raw}'"})
                skipped += 1
                continue
        else:
            month_key = None

        norm = cat_raw.lower()
        display_name.setdefault(norm, cat_raw)
        key = (norm, month_key)
        aggregated[key] = aggregated.get(key, 0.0) + limit_val

    conn = get_connection()
    cursor = conn.cursor()

    existing = {r["name"].lower(): r["name"]
                for r in conn.execute("SELECT name FROM expense_types WHERE user_id = %s", (user_id,)).fetchall()}
    used_colors = {r["color"] for r in conn.execute("SELECT color FROM expense_types WHERE user_id = %s", (user_id,)).fetchall()}

    new_categories = {norm for norm, _ in aggregated} - existing.keys()
    available_colors = [c for c in _DEFAULT_COLORS if c not in used_colors] or _DEFAULT_COLORS
    for idx, norm in enumerate(sorted(new_categories)):
        name = display_name[norm]
        color = available_colors[idx % len(available_colors)]
        sort_order = conn.execute("SELECT COALESCE(MAX(sort_order)+1, 0) AS sort_order FROM expense_types WHERE user_id = %s", (user_id,)).fetchone()["sort_order"]
        cursor.execute(
            "INSERT INTO expense_types (id, name, color, icon, sort_order, user_id) VALUES (%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), name, color, _DEFAULT_ICON, sort_order, user_id),
        )
        existing[norm] = name

    imported = 0
    for (norm, month_key), total in aggregated.items():
        type_name = existing[norm]
        if month_key:
            cursor.execute(
                "INSERT INTO monthly_budgets (user_id, type, month, monthly_limit) VALUES (%s,%s,%s,%s) ON CONFLICT (user_id, type, month) DO UPDATE SET monthly_limit = EXCLUDED.monthly_limit",
                (user_id, type_name, month_key, total),
            )
        else:
            cursor.execute(
                "INSERT INTO budgets (user_id, type, monthly_limit) VALUES (%s,%s,%s) ON CONFLICT (user_id, type) DO UPDATE SET monthly_limit = EXCLUDED.monthly_limit",
                (user_id, type_name, total),
            )
        imported += 1

    conn.commit()
    conn.close()
    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post("/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    header_row: Optional[int] = Form(None),
    sheet_name: Optional[str] = Form(None),
):
    content = await file.read()
    _check_file_size(content)
    lower = file.filename.lower()

    sheet_names = []
    if lower.endswith((".xlsx", ".xls")):
        sheet_names = _xlsx_sheet_names(content)

    raw_rows = _get_raw_rows(content, file.filename, sheet_name)
    detected = header_row if header_row is not None else _detect_header_row(raw_rows)
    headers, data_rows = _rows_to_dicts(raw_rows, detected)

    return {
        "headers": headers,
        "preview": data_rows[:3],
        "header_row": detected,
        "sheet_names": sheet_names,
    }


@router.post("/import")
async def import_records(
    file: UploadFile = File(...),
    mapping: str = Form(...),
    header_row: int = Form(...),
    sheet_name: Optional[str] = Form(None),
    confirmed_types: Optional[str] = Form(None),
    user_id: str = Depends(get_current_user),
):
    try:
        col_map = json.loads(mapping)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid mapping JSON")

    content = await file.read()
    _check_file_size(content)
    raw_rows = _get_raw_rows(content, file.filename, sheet_name)
    headers, data_rows = _rows_to_dicts(raw_rows, header_row)

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM expense_types WHERE user_id = %s", (user_id,))
    valid_types = [r["name"] for r in cursor.fetchall()]
    cursor.execute("SELECT pattern, expense_type FROM import_rules WHERE user_id = %s", (user_id,))
    rules = [(r["pattern"], r["expense_type"]) for r in cursor.fetchall()]

    imported = 0
    errors = []
    savings_expenses = []
    ai_categorized_count = 0
    from database import get_user_settings
    user_ai_settings = get_user_settings(conn, user_id)
    ai_categorize_fn = _get_ai_categorizer() if user_ai_settings["ai_enabled"] else None
    overrides = json.loads(confirmed_types) if confirmed_types else {}

    for i, row in enumerate(data_rows, start=header_row + 2):
        try:
            name = row.get(col_map.get("name", ""), "").strip()
            if not name:
                raise ValueError("Name is empty")

            raw_amt = row.get(col_map.get("amount", ""), "")
            amount_float = float(str(raw_amt).replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip())
            # float("nan")/float("inf") parse without raising ValueError, and NaN
            # fails every comparison (including == 0), so both would otherwise slip
            # past the zero-check below and corrupt SUM() aggregates downstream.
            if not math.isfinite(amount_float):
                raise ValueError("Amount is not a valid number")
            if amount_float == 0:
                raise ValueError("Amount is zero")
            if abs(amount_float) > 100_000_000:
                raise ValueError("Amount is unreasonably large")
            amount = round(abs(amount_float), 2)

            raw_date = row.get(col_map.get("date", ""), "").strip()
            date_str = _parse_date(raw_date)

            is_recurring = 0
            if col_map.get("is_recurring"):
                val = row.get(col_map["is_recurring"], "").strip().lower()
                is_recurring = 1 if val in ("1", "true", "yes", "y") else 0

            record_type_col_val = row.get(col_map.get("record_type", ""), "") if col_map.get("record_type") else ""
            row_type = _determine_record_type(amount_float, record_type_col_val)

            new_id = str(uuid.uuid4())
            now = datetime.now().isoformat()

            if row_type == "expense":
                row_0idx = str(i - (header_row + 2))
                confirmed = overrides.get(row_0idx)

                if confirmed and confirmed in valid_types:
                    expense_type = confirmed
                else:
                    raw_type = row.get(col_map.get("type", ""), "").strip() if col_map.get("type") else ""

                    if raw_type and raw_type.lower() == "savings":
                        if "Savings" not in valid_types:
                            cursor.execute(
                                "INSERT INTO expense_types (id, name, color, icon, sort_order, is_default, user_id) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (user_id, name) DO NOTHING",
                                (str(uuid.uuid4()), "Savings", "#8fb996", "Savings", 99, 1, user_id),
                            )
                            valid_types.append("Savings")
                        expense_type = "Savings"
                    elif raw_type:
                        expense_type = _soft_match_type(raw_type, valid_types) or _infer_type(name, valid_types, rules)
                    else:
                        expense_type = _infer_type(name, valid_types, rules)
                        # If the keyword matcher gave up and fell back to 'Other', escalate to AI.
                        # The LLM recognises brands and merchant names that keyword lists miss.
                        if ai_categorize_fn and _is_other_fallback(expense_type, valid_types):
                            suggestion = ai_categorize_fn(name, amount, valid_types)
                            suggested = suggestion.get("category")
                            if suggested and suggested in valid_types:
                                expense_type = suggested
                                ai_categorized_count += 1

                cursor.execute(
                    "INSERT INTO expenses (id, name, amount, type, date, created_at, is_recurring, user_id) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_id, name, amount, expense_type, date_str, now, is_recurring, user_id),
                )

                if expense_type == "Savings":
                    savings_expenses.append({"id": new_id, "name": name, "amount": amount, "date": date_str})
            else:
                cursor.execute(
                    "INSERT INTO incomes (id, name, amount, date, created_at, is_recurring, user_id) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (new_id, name, amount, date_str, now, is_recurring, user_id),
                )

            imported += 1

        except (ValueError, KeyError, TypeError) as e:
            errors.append({"row": i, "reason": str(e)})

    conn.commit()
    conn.close()

    return {
        "imported": imported,
        "skipped": len(errors),
        "errors": errors,
        "savings_expenses": savings_expenses,
        "ai_categorized": ai_categorized_count,
    }


@router.post("/import/legacy-db")
async def import_legacy_db(
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user),
):
    if not (file.filename or "").endswith(".db"):
        raise HTTPException(status_code=400, detail="File must be a .db SQLite database")

    content = await file.read()
    _check_file_size(content)

    def _col(row, key, default=None):
        try:
            return row[key]
        except IndexError:
            return default

    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    stats = {}
    try:
        old_db = _sqlite3.connect(tmp_path)
        old_db.row_factory = _sqlite3.Row
        old_cur = old_db.cursor()
        conn = get_connection()
        cursor = conn.cursor()
        now = datetime.now().isoformat()

        # macrocategories — dedup by name, track old→new ID for expense_types linking
        macro_id_map = {}
        try:
            old_cur.execute("SELECT * FROM macrocategories")
            count = 0
            for row in old_cur.fetchall():
                new_id = str(uuid.uuid4())
                cursor.execute(
                    "INSERT INTO macrocategories (id, name, color, budget_limit, user_id) "
                    "VALUES (%s,%s,%s,%s,%s) ON CONFLICT (user_id, name) DO NOTHING",
                    (new_id, row["name"], _col(row, "color", "#a0a0a0"), _col(row, "budget_limit"), user_id),
                )
                if cursor.rowcount > 0:
                    macro_id_map[row["id"]] = new_id
                    count += 1
                else:
                    existing = conn.execute(
                        "SELECT id FROM macrocategories WHERE name = %s AND user_id = %s", (row["name"], user_id)
                    ).fetchone()
                    if existing:
                        macro_id_map[row["id"]] = existing["id"]
            stats["macrocategories"] = count
        except Exception:
            stats["macrocategories"] = 0

        # expense_types — dedup by name, preserve macrocategory links
        try:
            old_cur.execute("SELECT * FROM expense_types")
            count = 0
            for row in old_cur.fetchall():
                old_macro = _col(row, "macrocategory_id")
                cursor.execute(
                    "INSERT INTO expense_types (id, name, color, icon, sort_order, is_default, user_id, macrocategory_id) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (user_id, name) DO NOTHING",
                    (str(uuid.uuid4()), row["name"], row["color"], row["icon"],
                     _col(row, "sort_order", 0), _col(row, "is_default", 0), user_id,
                     macro_id_map.get(old_macro) if old_macro else None),
                )
                if cursor.rowcount > 0:
                    count += 1
            stats["expense_types"] = count
        except Exception:
            stats["expense_types"] = 0

        # budgets — DO NOTHING on conflict (don't overwrite cloud budgets)
        try:
            old_cur.execute("SELECT * FROM budgets")
            count = 0
            for row in old_cur.fetchall():
                cursor.execute(
                    "INSERT INTO budgets (user_id, type, monthly_limit) VALUES (%s,%s,%s) "
                    "ON CONFLICT (user_id, type) DO NOTHING",
                    (user_id, row["type"], row["monthly_limit"]),
                )
                if cursor.rowcount > 0:
                    count += 1
            stats["budgets"] = count
        except Exception:
            stats["budgets"] = 0

        # monthly_budgets
        try:
            old_cur.execute("SELECT * FROM monthly_budgets")
            count = 0
            for row in old_cur.fetchall():
                cursor.execute(
                    "INSERT INTO monthly_budgets (user_id, type, month, monthly_limit) VALUES (%s,%s,%s,%s) "
                    "ON CONFLICT (user_id, type, month) DO NOTHING",
                    (user_id, row["type"], row["month"], row["monthly_limit"]),
                )
                if cursor.rowcount > 0:
                    count += 1
            stats["monthly_budgets"] = count
        except Exception:
            stats["monthly_budgets"] = 0

        # expenses — all get new UUIDs; keep old→new map for contribution linking
        exp_id_map = {}
        try:
            old_cur.execute("SELECT * FROM expenses")
            count = 0
            for row in old_cur.fetchall():
                new_id = str(uuid.uuid4())
                exp_id_map[row["id"]] = new_id
                cursor.execute(
                    "INSERT INTO expenses (id, name, amount, type, date, created_at, is_recurring, user_id) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_id, row["name"], row["amount"], row["type"], row["date"],
                     _col(row, "created_at") or now, _col(row, "is_recurring", 0), user_id),
                )
                count += 1
            stats["expenses"] = count
        except Exception:
            stats["expenses"] = 0

        # incomes — all get new UUIDs
        try:
            old_cur.execute("SELECT * FROM incomes")
            count = 0
            for row in old_cur.fetchall():
                cursor.execute(
                    "INSERT INTO incomes (id, name, amount, date, created_at, is_recurring, credit_type, user_id) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (str(uuid.uuid4()), row["name"], row["amount"], row["date"],
                     _col(row, "created_at") or now, _col(row, "is_recurring", 0),
                     _col(row, "credit_type"), user_id),
                )
                count += 1
            stats["incomes"] = count
        except Exception:
            stats["incomes"] = 0

        # import_rules — dedup by pattern
        try:
            old_cur.execute("SELECT * FROM import_rules")
            count = 0
            for row in old_cur.fetchall():
                cursor.execute(
                    "INSERT INTO import_rules (id, pattern, expense_type, created_at, user_id) "
                    "VALUES (%s,%s,%s,%s,%s) ON CONFLICT (user_id, pattern) DO NOTHING",
                    (str(uuid.uuid4()), row["pattern"], row["expense_type"],
                     _col(row, "created_at") or now, user_id),
                )
                if cursor.rowcount > 0:
                    count += 1
            stats["import_rules"] = count
        except Exception:
            stats["import_rules"] = 0

        # savings_goals — new UUIDs, track old→new for contributions
        goal_id_map = {}
        try:
            old_cur.execute("SELECT * FROM savings_goals")
            count = 0
            for row in old_cur.fetchall():
                new_id = str(uuid.uuid4())
                goal_id_map[row["id"]] = new_id
                cursor.execute(
                    "INSERT INTO savings_goals "
                    "(id, goal_type, name, target, deadline, created_at, color, allocation_pct, priority, paused, months_target, user_id) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (new_id, row["goal_type"], row["name"], row["target"],
                     _col(row, "deadline"), _col(row, "created_at") or now,
                     _col(row, "color"), _col(row, "allocation_pct"), _col(row, "priority"),
                     _col(row, "paused", 0), _col(row, "months_target"), user_id),
                )
                count += 1
            stats["savings_goals"] = count
        except Exception:
            stats["savings_goals"] = 0

        # savings_contributions — remap goal_id and expense_id to new UUIDs
        try:
            old_cur.execute("SELECT * FROM savings_contributions")
            count = 0
            for row in old_cur.fetchall():
                new_goal_id = goal_id_map.get(row["goal_id"])
                if not new_goal_id:
                    continue
                old_exp_id = _col(row, "expense_id")
                cursor.execute(
                    "INSERT INTO savings_contributions (id, goal_id, amount, date, note, created_at, expense_id, user_id) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (str(uuid.uuid4()), new_goal_id, row["amount"], row["date"],
                     _col(row, "note"), _col(row, "created_at") or now,
                     exp_id_map.get(old_exp_id) if old_exp_id else None, user_id),
                )
                count += 1
            stats["savings_contributions"] = count
        except Exception:
            stats["savings_contributions"] = 0

        conn.commit()
        conn.close()
        old_db.close()

    finally:
        os.unlink(tmp_path)

    return {"imported": stats}
